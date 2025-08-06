import os
import shutil
import time
import re
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import pyperclip
from datetime import datetime
import threading

class FileMonitorHandler(FileSystemEventHandler):
    """文件监控处理器"""
    def __init__(self, callback):
        self.callback = callback
        
    def on_created(self, event):
        if not event.is_directory:
            self.callback(f"创建文件: {event.src_path}")
    
    def on_deleted(self, event):
        if not event.is_directory:
            self.callback(f"删除文件: {event.src_path}")
    
    def on_modified(self, event):
        if not event.is_directory:
            self.callback(f"修改文件: {event.src_path}")
    
    def on_moved(self, event):
        if not event.is_directory:
            self.callback(f"移动文件: {event.src_path} -> {event.dest_path}")

class ClipboardHistory:
    """剪贴板历史管理"""
    def __init__(self, max_items=50):
        self.history = []
        self.max_items = max_items
        self.last_content = ""
        self.running = False
        self.thread = None
    
    def start_monitoring(self, callback=None):
        """开始监控剪贴板"""
        self.running = True
        self.callback = callback
        self.thread = threading.Thread(target=self._monitor_clipboard)
        self.thread.daemon = True
        self.thread.start()
    
    def stop_monitoring(self):
        """停止监控剪贴板"""
        self.running = False
        if self.thread:
            self.thread.join(timeout=1)
    
    def _monitor_clipboard(self):
        """监控剪贴板内容变化"""
        while self.running:
            try:
                current_content = pyperclip.paste()
                if current_content != self.last_content and current_content.strip():
                    self.add_to_history(current_content)
                    self.last_content = current_content
                    if self.callback:
                        self.callback(current_content)
                time.sleep(0.5)  # 每0.5秒检查一次
            except:
                pass
    
    def add_to_history(self, content):
        """添加到历史记录"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        item = {
            'content': content,
            'timestamp': timestamp,
            'type': self._detect_content_type(content)
        }
        
        # 避免重复
        if not self.history or self.history[0]['content'] != content:
            self.history.insert(0, item)
            if len(self.history) > self.max_items:
                self.history.pop()
    
    def _detect_content_type(self, content):
        """检测内容类型"""
        if re.match(r'^https?://', content):
            return '链接'
        elif re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', content):
            return '邮箱'
        elif content.isdigit():
            return '数字'
        elif len(content.split('\n')) > 1:
            return '多行文本'
        else:
            return '文本'
    
    def get_history(self):
        """获取历史记录"""
        return self.history.copy()
    
    def clear_history(self):
        """清空历史记录"""
        self.history.clear()

def create_excel_from_data(data, headers, output_path, sheet_name="Sheet1"):
    """从数据创建Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 添加表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 添加数据
    for row, row_data in enumerate(data, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    return output_path

def merge_excel_files(file_paths, output_path, merge_type='sheets'):
    """合并Excel文件"""
    if merge_type == 'sheets':
        # 合并为不同工作表
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for i, file_path in enumerate(file_paths):
                df = pd.read_excel(file_path)
                sheet_name = f"Sheet{i+1}"
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # 合并为同一工作表
        all_data = []
        for file_path in file_paths:
            df = pd.read_excel(file_path)
            all_data.append(df)
        
        combined_df = pd.concat(all_data, ignore_index=True)
        combined_df.to_excel(output_path, index=False)
    
    return output_path

def split_excel_file(file_path, output_dir, split_column=None, rows_per_file=1000):
    """拆分Excel文件"""
    df = pd.read_excel(file_path)
    output_files = []
    
    if split_column and split_column in df.columns:
        # 按列值拆分
        unique_values = df[split_column].unique()
        for value in unique_values:
            subset = df[df[split_column] == value]
            output_file = os.path.join(output_dir, f"{split_column}_{value}.xlsx")
            subset.to_excel(output_file, index=False)
            output_files.append(output_file)
    else:
        # 按行数拆分
        total_rows = len(df)
        num_files = (total_rows + rows_per_file - 1) // rows_per_file
        
        for i in range(num_files):
            start_row = i * rows_per_file
            end_row = min((i + 1) * rows_per_file, total_rows)
            subset = df.iloc[start_row:end_row]
            
            output_file = os.path.join(output_dir, f"part_{i+1}.xlsx")
            subset.to_excel(output_file, index=False)
            output_files.append(output_file)
    
    return output_files

def batch_rename_advanced(directory, pattern_type, pattern_data):
    """高级批量重命名"""
    if not os.path.exists(directory):
        raise ValueError("目录不存在")
    
    renamed_files = []
    files = [f for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f))]
    
    for i, filename in enumerate(files):
        old_path = os.path.join(directory, filename)
        name, ext = os.path.splitext(filename)
        
        if pattern_type == "sequence":
            # 序号重命名
            prefix = pattern_data.get('prefix', 'file')
            start_num = pattern_data.get('start_num', 1)
            digits = pattern_data.get('digits', 3)
            new_name = f"{prefix}{str(start_num + i).zfill(digits)}{ext}"
        
        elif pattern_type == "date":
            # 按文件修改时间重命名
            mtime = os.path.getmtime(old_path)
            date_format = pattern_data.get('date_format', '%Y%m%d_%H%M%S')
            prefix = pattern_data.get('prefix', '')
            suffix = pattern_data.get('suffix', '')
            
            date_str = datetime.fromtimestamp(mtime).strftime(date_format)
            new_name = f"{prefix}{date_str}{suffix}{ext}"
        
        elif pattern_type == "replace":
            # 替换重命名
            old_text = pattern_data.get('old_text', '')
            new_text = pattern_data.get('new_text', '')
            use_regex = pattern_data.get('use_regex', False)
            
            if use_regex:
                new_name = re.sub(old_text, new_text, name) + ext
            else:
                new_name = name.replace(old_text, new_text) + ext
        
        elif pattern_type == "case":
            # 大小写转换
            case_type = pattern_data.get('case_type', 'lower')
            if case_type == 'lower':
                new_name = filename.lower()
            elif case_type == 'upper':
                new_name = filename.upper()
            elif case_type == 'title':
                new_name = name.title() + ext
            else:
                new_name = filename
        
        else:
            continue
        
        new_path = os.path.join(directory, new_name)
        
        # 避免重名
        counter = 1
        while os.path.exists(new_path):
            name_part, ext_part = os.path.splitext(new_name)
            new_name = f"{name_part}_{counter}{ext_part}"
            new_path = os.path.join(directory, new_name)
            counter += 1
        
        try:
            os.rename(old_path, new_path)
            renamed_files.append((filename, new_name))
        except Exception as e:
            print(f"重命名失败 {filename}: {str(e)}")
    
    return renamed_files

def monitor_directory(directory, callback, include_subdirs=False):
    """监控目录变化"""
    if not os.path.exists(directory):
        raise ValueError("目录不存在")
    
    event_handler = FileMonitorHandler(callback)
    observer = Observer()
    observer.schedule(event_handler, directory, recursive=include_subdirs)
    observer.start()
    
    return observer

def convert_file_format(input_file, output_file, format_type):
    """转换文件格式"""
    if format_type == "excel_to_csv":
        df = pd.read_excel(input_file)
        df.to_csv(output_file, index=False, encoding='utf-8-sig')
    
    elif format_type == "csv_to_excel":
        df = pd.read_csv(input_file, encoding='utf-8-sig')
        df.to_excel(output_file, index=False)
    
    elif format_type == "excel_to_json":
        df = pd.read_excel(input_file)
        df.to_json(output_file, orient='records', force_ascii=False, indent=2)
    
    elif format_type == "json_to_excel":
        df = pd.read_json(input_file)
        df.to_excel(output_file, index=False)
    
    else:
        raise ValueError(f"不支持的转换类型: {format_type}")
    
    return output_file

def clean_excel_data(input_file, output_file, operations):
    """清理Excel数据"""
    df = pd.read_excel(input_file)
    
    for operation in operations:
        if operation == "remove_duplicates":
            df = df.drop_duplicates()
        
        elif operation == "remove_empty_rows":
            df = df.dropna(how='all')
        
        elif operation == "remove_empty_columns":
            df = df.dropna(axis=1, how='all')
        
        elif operation == "trim_whitespace":
            for col in df.select_dtypes(include=['object']):
                df[col] = df[col].astype(str).str.strip()
        
        elif operation == "standardize_case":
            for col in df.select_dtypes(include=['object']):
                df[col] = df[col].astype(str).str.title()
    
    df.to_excel(output_file, index=False)
    return output_file

def analyze_directory_structure(directory):
    """分析目录结构"""
    if not os.path.exists(directory):
        raise ValueError("目录不存在")
    
    total_size = 0
    file_count = 0
    dir_count = 0
    file_types = {}
    large_files = []
    
    for root, dirs, files in os.walk(directory):
        dir_count += len(dirs)
        
        for file in files:
            file_path = os.path.join(root, file)
            try:
                size = os.path.getsize(file_path)
                total_size += size
                file_count += 1
                
                # 统计文件类型
                _, ext = os.path.splitext(file)
                ext = ext.lower()
                if ext in file_types:
                    file_types[ext]['count'] += 1
                    file_types[ext]['size'] += size
                else:
                    file_types[ext] = {'count': 1, 'size': size}
                
                # 记录大文件（大于10MB）
                if size > 10 * 1024 * 1024:
                    large_files.append({
                        'path': file_path,
                        'size': size,
                        'size_mb': round(size / (1024 * 1024), 2)
                    })
                
            except OSError:
                pass
    
    return {
        'total_size': total_size,
        'total_size_mb': round(total_size / (1024 * 1024), 2),
        'file_count': file_count,
        'dir_count': dir_count,
        'file_types': file_types,
        'large_files': large_files
    }